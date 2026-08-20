"""Write DB schema back to Excel design document."""

from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import openpyxl

from .excel_parser import (
    HEADER_ROW,
    LABEL_TO_FIELD,
    _cell,
    _effective_cell,
    _find_flat_header,
    _is_block_column_header,
    _is_table_meta_row,
    _map_header_row,
    _split_index_key_text,
    normalize_label,
    resolve_sheet,
)
from .schema_reader import DbTableInfo
from .type_mapper import map_pg_to_excel
from openpyxl.styles import Alignment, Border, Side
from openpyxl.styles.borders import BORDER_THIN

SHEET_NAME = "테이블정의서"

HEADER_LABELS = [
    "No",
    "DB명",
    "스키마명",
    "한글 테이블명",
    "영문 테이블명",
    "한글 컬럼명",
    "영문 컬럼명",
    "데이터 타입",
    "데이터 길이",
    "Not Null 여부",
    "PK 여부",
    "FK 여부",
    "암호화여부",
    "Index Key",
    "개인정보",
    "코멘트",
]


@dataclass
class TemplateLayout:
    format: str  # flat | block
    sheet_name: str
    header_row: int | None
    col_map: dict[str, int]
    has_db: bool
    has_schema: bool
    has_module_row: bool = False


@dataclass
class FlatPrototype:
    preamble_rows: list[dict[int, object]]
    header_values: dict[int, object]
    header_row: int


@dataclass
class BlockIndexKeyFooter:
    label_col: int
    label: str
    value_col: int
    label_end_col: int
    value_end_col: int


@dataclass
class BlockPrototype:
    has_module_row: bool
    module_col1: object | None
    module_col3: object | None
    meta_col1: object | None
    meta_col5: object | None
    header_values: dict[int, object]
    col_map: dict[str, int]
    index_key_footer: BlockIndexKeyFooter | None = None


def write_schema_to_excel_bytes(
    tables: list[DbTableInfo],
    *,
    db_name: str = "dbm",
    schema_name: str = "db1",
    template: Path | BinaryIO | None = None,
    sheet_name: str | None = SHEET_NAME,
    system_name: str = "",
    created_date: str = "",
    author: str = "",
) -> bytes:
    """Write tables into the template workbook, keeping merges, sizes, and borders."""
    db_label = (db_name or "dbm").strip() or "dbm"
    schema_label = (schema_name or "db1").strip() or "db1"
    system_label = (system_name or "").strip()
    created_label = (created_date or "").strip()
    author_label = (author or "").strip()

    if template is not None:
        wb = openpyxl.load_workbook(template)
        try:
            layout = _analyze_workbook(wb, sheet_name)
            ws = wb[layout.sheet_name]
            if layout.format == "block":
                _fill_block_template(
                    ws,
                    layout,
                    tables,
                    db_label,
                    schema_label,
                    system_label,
                    created_label,
                    author_label,
                )
            else:
                _fill_flat_template(
                    ws,
                    layout,
                    tables,
                    db_label,
                    schema_label,
                    system_label,
                    created_label,
                    author_label,
                )
            buf = BytesIO()
            wb.save(buf)
            return buf.getvalue()
        finally:
            wb.close()

    wb_default = _new_workbook(sheet_name or SHEET_NAME)
    ws_default = wb_default[wb_default.sheetnames[0]]
    layout = TemplateLayout(
        format="flat",
        sheet_name=ws_default.title,
        header_row=HEADER_ROW,
        col_map=_map_header_row(ws_default, HEADER_ROW),
        has_db=True,
        has_schema=True,
    )
    flat_proto = FlatPrototype(
        preamble_rows=[{1: "테이블정의서"}],
        header_values={col: label for col, label in enumerate(HEADER_LABELS, start=1)},
        header_row=HEADER_ROW,
    )
    _write_flat_fresh(ws_default, layout, flat_proto, tables, db_label, schema_label)
    _fill_doc_meta_area(
        ws_default, 1, HEADER_ROW, 8, system_label, created_label, author_label
    )
    buf = BytesIO()
    wb_default.save(buf)
    wb_default.close()
    return buf.getvalue()


def default_export_filename(schema: str) -> str:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in schema) or "schema"
    return f"design_{safe}_{stamp}.xlsx"


def _read_template_layout(
    template: Path | BinaryIO,
    sheet_name: str | None,
) -> tuple[TemplateLayout, FlatPrototype | None, BlockPrototype | None]:
    wb = openpyxl.load_workbook(template, data_only=True)
    try:
        layout = _analyze_workbook(wb, sheet_name)
        ws = wb[layout.sheet_name]
        if layout.format == "block":
            return layout, None, _read_block_prototype(ws, layout)
        return layout, _read_flat_prototype(ws, layout), None
    finally:
        wb.close()


def _analyze_workbook(wb, sheet_name: str | None) -> TemplateLayout:
    resolved = resolve_sheet(wb, sheet_name)
    ws = wb[resolved]
    flat = _find_flat_header(ws)
    if flat:
        header_row, col_map = flat
        return TemplateLayout(
            format="flat",
            sheet_name=resolved,
            header_row=header_row,
            col_map=col_map,
            has_db="db" in col_map,
            has_schema="schema" in col_map,
        )

    max_row = ws.max_row or 0
    for row in range(1, max_row + 1):
        if _is_block_column_header(ws, row):
            meta_row = row - 1
            has_module = (
                meta_row > 1
                and str(_cell(ws, meta_row - 1, 1) or "").strip().startswith("모듈")
            )
            col_map = _map_header_row(ws, row)
            return TemplateLayout(
                format="block",
                sheet_name=resolved,
                header_row=row,
                col_map=col_map,
                has_db="db" in col_map,
                has_schema="schema" in col_map,
                has_module_row=has_module,
            )
    raise ValueError(
        f"시트 '{resolved}'에서 테이블정의서 양식 헤더를 찾지 못했습니다."
    )


def _read_flat_prototype(ws, layout: TemplateLayout) -> FlatPrototype:
    header_row = layout.header_row or HEADER_ROW
    preamble_rows: list[dict[int, object]] = []
    for row in range(1, header_row):
        values: dict[int, object] = {}
        for col in range(1, (ws.max_column or 0) + 1):
            val = _cell(ws, row, col)
            if val is not None and str(val).strip():
                values[col] = val
        if values:
            preamble_rows.append(values)

    header_values: dict[int, object] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        val = _cell(ws, header_row, col)
        if val is not None:
            header_values[col] = val

    return FlatPrototype(
        preamble_rows=preamble_rows,
        header_values=header_values,
        header_row=header_row,
    )


def _read_block_prototype(ws, layout: TemplateLayout) -> BlockPrototype:
    header_row = layout.header_row or 1
    meta_row = header_row - 1
    module_row = meta_row - 1 if layout.has_module_row else None

    header_values: dict[int, object] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        val = _cell(ws, header_row, col)
        if val is not None:
            header_values[col] = val

    return BlockPrototype(
        has_module_row=layout.has_module_row,
        module_col1=_cell(ws, module_row, 1) if module_row else None,
        module_col3=_cell(ws, module_row, 3) if module_row else None,
        meta_col1=_cell(ws, meta_row, 1),
        meta_col5=_cell(ws, meta_row, 5),
        header_values=header_values,
        col_map=layout.col_map,
        index_key_footer=_read_block_index_key_footer(ws, header_row),
    )


def _read_block_index_key_footer(ws, header_row: int) -> BlockIndexKeyFooter | None:
    """Block templates may store Index Key in a footer row, not a header column."""
    max_row = ws.max_row or 0
    data_row = header_row + 1
    while data_row <= max_row:
        if _is_table_meta_row(ws, data_row):
            break
        marker = str(_cell(ws, data_row, 1) or "").strip()
        if marker == "Index Key":
            max_col = max(ws.max_column or 8, 8)
            label_col, label_end = _row_merge_span(ws, data_row, 1)
            value_col = 3
            for col in range(label_end + 1, max_col + 1):
                if _cell(ws, data_row, col) not in (None, ""):
                    value_col = col
                    break
            else:
                value_col = max(label_end + 1, 3)
            value_start, value_end = _row_merge_span(ws, data_row, value_col)
            if label_end <= label_col and value_end <= value_start:
                label_col, label_end = 1, min(2, value_col - 1) if value_col > 1 else 1
                if label_end < 1:
                    label_end = 1
                value_col, value_end = max(label_end + 1, 3), max_col
            elif value_end <= value_start:
                value_end = max_col
            if label_end < label_col:
                label_end = label_col
            if value_end < value_col:
                value_end = max(value_col, max_col)
            return BlockIndexKeyFooter(
                label_col=label_col,
                label=marker,
                value_col=value_col,
                label_end_col=label_end,
                value_end_col=value_end,
            )
        if marker in ("업무규칙", "테이블 정의서", "테이블정의서"):
            break
        column_name = _cell(ws, data_row, 2)
        data_type = _cell(ws, data_row, 5)
        if column_name and data_type:
            data_row += 1
            continue
        data_row += 1
    return None


def _collect_table_index_keys(table: DbTableInfo) -> list[str]:
    seen: set[str] = set()
    lines: list[str] = []

    def add(label: str) -> None:
        text = (label or "").strip()
        if not text or text in seen:
            return
        seen.add(text)
        lines.append(text)

    for col in table.columns:
        for part in _split_index_key_text(col.index_key or ""):
            add(part)

    pk_cols = [c.column_name.upper() for c in table.columns if c.is_pk]
    if pk_cols and not any(x.upper().startswith("PK_") for x in lines):
        add(f"PK_{table.name.upper()}({', '.join(pk_cols)})")
    for col in table.columns:
        if not col.is_fk:
            continue
        col_u = col.column_name.upper()
        if any(x.upper().startswith("FK_") and col_u in x.upper() for x in lines):
            continue
        parent = ""
        if col.fk_ref:
            parent = str(col.fk_ref).split("(", 1)[0].strip().upper()
        suffix = f"_{parent}" if parent else ""
        add(f"FK_{table.name.upper()}{suffix}({col_u})")

    def _priority(label: str) -> tuple[int, str]:
        upper = label.upper()
        if upper.startswith("PK_"):
            return (0, upper)
        if upper.startswith("UK_"):
            return (1, upper)
        if upper.startswith("FK_"):
            return (2, upper)
        return (3, upper)

    return sorted(lines, key=_priority)


@dataclass
class _CellSnap:
    value: object
    font: object | None
    border: object | None
    fill: object | None
    number_format: str
    protection: object | None
    alignment: object | None


@dataclass
class _RowSnap:
    height: float | None
    cells: list[_CellSnap]


def _snap_cell(cell) -> _CellSnap:
    return _CellSnap(
        value=cell.value,
        font=copy(cell.font) if cell.has_style else None,
        border=copy(cell.border) if cell.has_style else None,
        fill=copy(cell.fill) if cell.has_style else None,
        number_format=cell.number_format,
        protection=copy(cell.protection) if cell.has_style else None,
        alignment=copy(cell.alignment) if cell.has_style else None,
    )


def _apply_cell(cell, snap: _CellSnap, *, keep_value: bool = False) -> None:
    if not keep_value:
        try:
            cell.value = snap.value
        except AttributeError:
            pass
    if snap.font is not None:
        cell.font = copy(snap.font)
    if snap.border is not None:
        cell.border = copy(snap.border)
    if snap.fill is not None:
        cell.fill = copy(snap.fill)
    try:
        cell.number_format = snap.number_format
    except AttributeError:
        pass
    if snap.protection is not None:
        cell.protection = copy(snap.protection)
    if snap.alignment is not None:
        cell.alignment = copy(snap.alignment)


def _snapshot_row(ws, row: int, max_col: int) -> _RowSnap:
    height = None
    dim = ws.row_dimensions.get(row)
    if dim is not None:
        height = dim.height
    cells = [_snap_cell(ws.cell(row, col)) for col in range(1, max_col + 1)]
    return _RowSnap(height=height, cells=cells)


def _apply_row_snapshot(ws, row: int, snap: _RowSnap, *, clear_values: bool = False) -> None:
    if snap.height:
        ws.row_dimensions[row].height = snap.height
    _unmerge_overlap(ws, row, 1, row, max(len(snap.cells), 1))
    for idx, cell_snap in enumerate(snap.cells, start=1):
        cell = ws.cell(row, idx)
        _apply_cell(cell, cell_snap)
        if clear_values:
            cell.value = None


def _unmerge_overlap(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        overlap = not (
            rng.max_row < min_row
            or rng.min_row > max_row
            or rng.max_col < min_col
            or rng.min_col > max_col
        )
        if overlap:
            ws.unmerge_cells(str(rng))


def _snapshot_merges(ws, start_row: int, end_row: int) -> list[tuple[int, int, int, int]]:
    result: list[tuple[int, int, int, int]] = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start_row and rng.max_row <= end_row:
            result.append(
                (
                    rng.min_row - start_row,
                    rng.min_col,
                    rng.max_row - start_row,
                    rng.max_col,
                )
            )
    return result


def _paste_merges(ws, dest_row: int, merges: list[tuple[int, int, int, int]]) -> None:
    for rel_r1, c1, rel_r2, c2 in merges:
        ws.merge_cells(
            start_row=dest_row + rel_r1,
            start_column=c1,
            end_row=dest_row + rel_r2,
            end_column=c2,
        )


def _unmerge_from_row(ws, start_row: int) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= start_row:
            ws.unmerge_cells(str(rng))


def _safe_merge(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    if min_row > max_row or min_col > max_col:
        return
    if min_row == max_row and min_col == max_col:
        return
    for rng in list(ws.merged_cells.ranges):
        overlap = not (
            rng.max_row < min_row
            or rng.min_row > max_row
            or rng.max_col < min_col
            or rng.min_col > max_col
        )
        if overlap:
            ws.unmerge_cells(str(rng))
    ws.merge_cells(
        start_row=min_row,
        start_column=min_col,
        end_row=max_row,
        end_column=max_col,
    )


def _row_merge_span(ws, row: int, col: int) -> tuple[int, int]:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_col, rng.max_col
    return col, col


def _merge_index_key_row(ws, row: int, max_col: int, footer: BlockIndexKeyFooter) -> None:
    label_end = max(footer.label_col, footer.label_end_col)
    value_end = max(footer.value_col, footer.value_end_col, max_col)
    _safe_merge(ws, row, footer.label_col, row, label_end)
    _safe_merge(ws, row, footer.value_col, row, value_end)
    ws.cell(row, footer.label_col).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    ws.cell(row, footer.value_col).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )


def _write_index_key_row(
    ws,
    row: int,
    max_col: int,
    footer: BlockIndexKeyFooter,
    index_label: str,
    proto: "_RowSnap | None" = None,
    *,
    write_label: bool = True,
) -> None:
    if proto:
        _apply_row_snapshot(ws, row, proto, clear_values=True)
    else:
        for col in range(1, max_col + 1):
            ws.cell(row, col).value = None
    if write_label:
        ws.cell(row, footer.label_col).value = footer.label
    ws.cell(row, footer.value_col).value = index_label or None
    label_end = max(footer.label_col, footer.label_end_col)
    value_end = max(footer.value_col, footer.value_end_col, max_col)
    if write_label:
        _safe_merge(ws, row, footer.label_col, row, label_end)
    _safe_merge(ws, row, footer.value_col, row, value_end)
    ws.cell(row, footer.value_col).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )


def _write_index_key_block(
    ws,
    start_row: int,
    max_col: int,
    footer: BlockIndexKeyFooter,
    index_labels: list[str],
    proto: "_RowSnap | None" = None,
) -> int:
    keys = [k for k in index_labels if k]
    row_count = max(len(keys), 2)
    end_row = start_row + row_count - 1
    label_col = footer.label_col or 1
    value_col = max(footer.value_col, 2)
    label_end = max(footer.label_end_col, label_col)
    if value_col <= label_end:
        value_col = label_end + 1
    if label_end == label_col and value_col >= 3:
        label_end = min(2, value_col - 1)
        value_col = max(value_col, label_end + 1)
    value_end = max(value_col, max_col)

    _clip_rows_to_col(ws, start_row, end_row, value_end)
    for i in range(row_count):
        row = start_row + i
        if proto:
            _apply_row_snapshot(ws, row, proto, clear_values=True)
        else:
            for col in range(1, value_end + 1):
                ws.cell(row, col).value = None
        _clip_rows_to_col(ws, row, row, value_end)

    _clear_borders(ws, start_row, 1, end_row, value_end)
    box_side = _sample_border_side(proto)
    _apply_box_border(ws, start_row, label_col, end_row, label_end, box_side)
    _apply_box_border(ws, start_row, value_col, end_row, value_end, box_side)

    _safe_merge(ws, start_row, label_col, end_row, label_end)
    _write_cell(ws, start_row, label_col, footer.label)
    ws.cell(start_row, label_col).alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )

    _safe_merge(ws, start_row, value_col, end_row, value_end)
    _write_cell(ws, start_row, value_col, "\n".join(keys) or None)
    ws.cell(start_row, value_col).alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )

    line_h = 18
    if proto and proto.height:
        line_h = max(float(proto.height), 18)
    extra_lines = max(len(keys) - 1, 0)
    for row in range(start_row, end_row + 1):
        ws.row_dimensions[row].height = line_h + (extra_lines * 4)
    _clip_rows_to_col(ws, start_row, end_row, value_end)
    return end_row + 1


def _sample_border_side(proto: "_RowSnap | None") -> Side:
    if proto:
        for cell in proto.cells:
            border = cell.border
            if border is None:
                continue
            for attr in ("left", "right", "top", "bottom"):
                side = getattr(border, attr, None)
                if side is not None and side.style:
                    return Side(style=side.style, color=side.color)
    return Side(style=BORDER_THIN, color="000000")


_NO_SIDE = Side(border_style=None)
_NO_BORDER = Border(left=_NO_SIDE, right=_NO_SIDE, top=_NO_SIDE, bottom=_NO_SIDE)


def _set_cell_border(cell, border: Border) -> None:
    try:
        cell.border = border
    except AttributeError:
        pass


def _clear_borders(ws, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            _set_cell_border(ws.cell(row, col), _NO_BORDER)


def _apply_box_border(ws, min_row: int, min_col: int, max_row: int, max_col: int, side: Side) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            _set_cell_border(
                ws.cell(row, col),
                Border(
                    left=side if col == min_col else _NO_SIDE,
                    right=side if col == max_col else _NO_SIDE,
                    top=side if row == min_row else _NO_SIDE,
                    bottom=side if row == max_row else _NO_SIDE,
                ),
            )


def _table_grid_end_col(ws, header_row: int, col_map: dict[str, int], max_scan: int) -> int:
    last = max(col_map.values(), default=1)
    skip = _SYSTEM_LABELS | _DATE_LABELS | _AUTHOR_LABELS
    for col in range(1, max_scan + 1):
        label = normalize_label(_effective_cell(ws, header_row, col))
        if not label or label in skip:
            continue
        if label in LABEL_TO_FIELD:
            last = max(last, col)
    return last


def _clip_rows_to_col(ws, start_row: int, end_row: int, last_col: int, extra: int = 24) -> None:
    for rng in list(ws.merged_cells.ranges):
        if rng.max_row < start_row or rng.min_row > end_row:
            continue
        if rng.max_col > last_col:
            ws.unmerge_cells(str(rng))
    _clear_beyond_col(ws, start_row, end_row, last_col, extra)


def _clear_beyond_col(ws, start_row: int, end_row: int, last_col: int, extra: int = 16) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(last_col + 1, last_col + extra + 1):
            cell = ws.cell(row, col)
            try:
                cell.value = None
            except AttributeError:
                continue
            _set_cell_border(cell, _NO_BORDER)


def _sheet_max_col(ws, layout: TemplateLayout) -> int:
    mapped = max(layout.col_map.values(), default=1)
    return max(ws.max_column or 1, mapped, 8)


_SYSTEM_LABELS = {"시스템명", "모듈시스템명"}
_DATE_LABELS = {"작성일", "작성일자"}
_AUTHOR_LABELS = {"작성자", "작성자명"}


def _fill_labeled_value(
    ws, row_from: int, row_to: int, max_col: int, labels: set[str], value: str
) -> bool:
    if row_to < row_from:
        return False
    for row in range(row_from, row_to + 1):
        for col in range(1, max_col + 1):
            if normalize_label(_effective_cell(ws, row, col)) not in labels:
                continue
            _, label_end = _row_merge_span(ws, row, col)
            target_col = label_end + 1
            for c in range(label_end + 1, max_col + 1):
                origin = _merge_origin(ws, row, c)
                if origin and origin != (row, c):
                    continue
                label = normalize_label(_effective_cell(ws, row, c))
                if label in labels or label in _SYSTEM_LABELS | _DATE_LABELS | _AUTHOR_LABELS:
                    continue
                target_col = origin[1] if origin else c
                break
            if target_col <= max_col:
                _write_cell(ws, row, target_col, value or None)
            return True
    return False


def _write_doc_meta_row(
    ws, row: int, system_name: str, created_date: str, author: str
) -> None:
    ws.cell(row, 1).value = "시스템명"
    ws.cell(row, 3).value = system_name or None
    ws.cell(row, 5).value = "작성일"
    ws.cell(row, 6).value = created_date or None
    ws.cell(row, 7).value = "작성자"
    ws.cell(row, 8).value = author or None
    _safe_merge(ws, row, 1, row, 2)
    _safe_merge(ws, row, 3, row, 4)
    ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 5).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row, 7).alignment = Alignment(horizontal="center", vertical="center")


def _fill_doc_meta_area(
    ws,
    row_from: int,
    row_to: int,
    max_col: int,
    system_name: str,
    created_date: str,
    author: str,
) -> tuple[bool, bool]:
    _fill_labeled_value(ws, row_from, row_to, max_col, _SYSTEM_LABELS, system_name)
    found_date = _fill_labeled_value(
        ws, row_from, row_to, max_col, _DATE_LABELS, created_date
    )
    found_author = _fill_labeled_value(
        ws, row_from, row_to, max_col, _AUTHOR_LABELS, author
    )
    return found_date, found_author


def _snap_has_labels(snaps: list[_RowSnap], labels: set[str]) -> bool:
    for snap in snaps:
        for cell in snap.cells:
            if normalize_label(cell.value) in labels:
                return True
    return False


def _fill_system_name_area(
    ws, row_from: int, row_to: int, max_col: int, system_name: str
) -> None:
    _fill_labeled_value(ws, row_from, row_to, max_col, _SYSTEM_LABELS, system_name)


def _merge_origin(ws, row: int, col: int) -> tuple[int, int] | None:
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row, rng.min_col
    return None


def _write_cell(ws, row: int, col: int | None, value) -> None:
    if not col:
        return
    origin = _merge_origin(ws, row, col)
    r, c = origin if origin else (row, col)
    ws.cell(r, c).value = value


def _fill_flat_template(
    ws,
    layout: TemplateLayout,
    tables: list[DbTableInfo],
    db_name: str,
    schema_name: str,
    system_name: str,
    created_date: str = "",
    author: str = "",
) -> None:
    header_row = layout.header_row or HEADER_ROW
    max_col = _sheet_max_col(ws, layout)
    proto_row = header_row + 1
    if proto_row <= (ws.max_row or 0):
        proto = _snapshot_row(ws, proto_row, max_col)
    else:
        proto = _snapshot_row(ws, header_row, max_col)
    _fill_doc_meta_area(
        ws, 1, header_row, max_col, system_name, created_date, author
    )
    _unmerge_from_row(ws, header_row + 1)
    last = ws.max_row or header_row
    if last > header_row:
        ws.delete_rows(header_row + 1, last - header_row)

    row = header_row + 1
    next_no = 1
    for table in sorted(tables, key=lambda t: t.name):
        for col in table.columns:
            _apply_row_snapshot(ws, row, proto)
            excel_type, length = map_pg_to_excel(col.pg_type, col.max_length)
            if "no" in layout.col_map:
                ws.cell(row, layout.col_map["no"]).value = next_no
                next_no += 1
            _write_column_cells(
                ws,
                row,
                layout.col_map,
                layout,
                table,
                col,
                db_name,
                schema_name,
                excel_type,
                length,
            )
            row += 1


def _fill_block_template(
    ws,
    layout: TemplateLayout,
    tables: list[DbTableInfo],
    db_name: str,
    schema_name: str,
    system_name: str,
    created_date: str = "",
    author: str = "",
) -> None:
    max_row = ws.max_row or 0
    meta_row = None
    for row in range(1, max_row + 1):
        if _is_table_meta_row(ws, row):
            meta_row = row
            break
    if not meta_row:
        raise ValueError("블록형 양식에서 테이블명 행을 찾지 못했습니다.")

    has_module = layout.has_module_row
    block_start = meta_row - 1 if has_module and meta_row > 1 else meta_row
    header_row = meta_row + 1
    if header_row > max_row or not _is_block_column_header(ws, header_row):
        raise ValueError("블록형 양식에서 컬럼 헤더를 찾지 못했습니다.")

    meta_max_col = max(_sheet_max_col(ws, layout), 12)
    table_end_col = _table_grid_end_col(ws, header_row, layout.col_map, meta_max_col)
    header_snaps = [
        _snapshot_row(ws, r, meta_max_col) for r in range(block_start, header_row + 1)
    ]
    header_merges = _snapshot_merges(ws, block_start, header_row)
    data_proto = _snapshot_row(ws, header_row + 1, table_end_col)
    footer = _read_block_index_key_footer(ws, header_row)
    footer_proto = None
    if footer:
        footer_row = header_row + 1
        while footer_row <= max_row:
            marker = str(_cell(ws, footer_row, 1) or "").strip()
            if marker == "Index Key":
                footer_proto = _snapshot_row(ws, footer_row, table_end_col)
                break
            if _is_table_meta_row(ws, footer_row):
                break
            footer_row += 1

    keep_until = block_start - 1
    header_has_date = _snap_has_labels(header_snaps, _DATE_LABELS)
    header_has_author = _snap_has_labels(header_snaps, _AUTHOR_LABELS)
    found_date, found_author = _fill_doc_meta_area(
        ws, 1, keep_until, meta_max_col, system_name, created_date, author
    )
    _unmerge_from_row(ws, block_start)
    last = ws.max_row or block_start
    if last >= block_start:
        ws.delete_rows(block_start, last - block_start + 1)

    dest = block_start
    if (not found_date and not header_has_date) or (
        not found_author and not header_has_author
    ):
        _write_doc_meta_row(ws, dest, system_name, created_date, author)
        dest += 1
        ws.row_dimensions[dest].height = 12
        dest += 1
    col_map = layout.col_map
    for table in sorted(tables, key=lambda t: t.name):
        header_at = dest
        for i, snap in enumerate(header_snaps):
            _apply_row_snapshot(ws, dest + i, snap)
        _paste_merges(ws, dest, header_merges)
        if has_module:
            if system_name:
                ws.cell(dest, 3).value = system_name
            meta_at = dest + 1
        else:
            meta_at = dest
        ws.cell(meta_at, 3).value = table.korean_name
        ws.cell(meta_at, 6).value = table.name.upper()
        _fill_doc_meta_area(
            ws,
            header_at,
            header_at + len(header_snaps) - 1,
            meta_max_col,
            system_name,
            created_date,
            author,
        )
        dest = dest + len(header_snaps)
        row_no = 1
        for col in table.columns:
            _apply_row_snapshot(ws, dest, data_proto, clear_values=True)
            excel_type, length = map_pg_to_excel(col.pg_type, col.max_length)
            if "no" in col_map:
                ws.cell(dest, col_map["no"]).value = row_no
                row_no += 1
            _write_column_cells(
                ws,
                dest,
                col_map,
                layout,
                table,
                col,
                db_name,
                schema_name,
                excel_type,
                length,
            )
            dest += 1
        if footer:
            dest = _write_index_key_block(
                ws,
                dest,
                table_end_col,
                footer,
                _collect_table_index_keys(table),
                footer_proto,
            )
        ws.row_dimensions[dest].height = 18
        dest += 1


def _write_flat_fresh(
    ws,
    layout: TemplateLayout,
    prototype: FlatPrototype,
    tables: list[DbTableInfo],
    db_name: str,
    schema_name: str,
) -> None:
    row = 1
    for preamble in prototype.preamble_rows:
        for col_idx, value in preamble.items():
            ws.cell(row, col_idx).value = value
        row += 1

    for col_idx, value in prototype.header_values.items():
        ws.cell(row, col_idx).value = value
    col_map = layout.col_map
    row += 1

    next_no = 1
    for table in sorted(tables, key=lambda t: t.name):
        for col in table.columns:
            excel_type, length = map_pg_to_excel(col.pg_type, col.max_length)
            if "no" in col_map:
                ws.cell(row, col_map["no"]).value = next_no
                next_no += 1
            _write_column_cells(
                ws,
                row,
                col_map,
                layout,
                table,
                col,
                db_name,
                schema_name,
                excel_type,
                length,
            )
            row += 1


def _write_block_fresh(
    ws,
    layout: TemplateLayout,
    prototype: BlockPrototype,
    tables: list[DbTableInfo],
    db_name: str,
    schema_name: str,
) -> None:
    row = 1
    for table in sorted(tables, key=lambda t: t.name):
        if prototype.has_module_row:
            if prototype.module_col1 is not None:
                ws.cell(row, 1).value = prototype.module_col1
            if prototype.module_col3 is not None:
                ws.cell(row, 3).value = prototype.module_col3
            row += 1

        if prototype.meta_col1 is not None:
            ws.cell(row, 1).value = prototype.meta_col1
        if prototype.meta_col5 is not None:
            ws.cell(row, 5).value = prototype.meta_col5
        ws.cell(row, 3).value = table.korean_name
        ws.cell(row, 6).value = table.name.upper()
        row += 1

        for col_idx, value in prototype.header_values.items():
            ws.cell(row, col_idx).value = value
        col_map = prototype.col_map
        row += 1

        row_no = 1
        for col in table.columns:
            excel_type, length = map_pg_to_excel(col.pg_type, col.max_length)
            if "no" in col_map:
                ws.cell(row, col_map["no"]).value = row_no
                row_no += 1
            _write_column_cells(
                ws,
                row,
                col_map,
                layout,
                table,
                col,
                db_name,
                schema_name,
                excel_type,
                length,
            )
            row += 1

        if prototype.index_key_footer:
            footer = prototype.index_key_footer
            row = _write_index_key_block(
                ws, row, 8, footer, _collect_table_index_keys(table)
            )
        ws.row_dimensions[row].height = 18
        row += 1


def _write_column_cells(
    ws,
    row: int,
    col_map: dict[str, int],
    layout: TemplateLayout,
    table: DbTableInfo,
    col,
    db_name: str,
    schema_name: str,
    excel_type: str,
    length,
) -> None:
    if layout.has_db and "db" in col_map:
        _write_cell(ws, row, col_map["db"], db_name.upper())
    if layout.has_schema and "schema" in col_map:
        _write_cell(ws, row, col_map["schema"], schema_name)

    if "table_ko" in col_map:
        _write_cell(ws, row, col_map["table_ko"], table.korean_name)
    if "table_en" in col_map:
        _write_cell(ws, row, col_map["table_en"], table.name.upper())
    if "column_ko" in col_map:
        _write_cell(
            ws, row, col_map["column_ko"], col.korean_name or col.column_name
        )
    if "column_en" in col_map:
        _write_cell(ws, row, col_map["column_en"], col.column_name.upper())
    if "data_type" in col_map:
        _write_cell(ws, row, col_map["data_type"], excel_type)
    if "data_length" in col_map:
        length_val = int(length) if isinstance(length, float) else length
        _write_cell(ws, row, col_map["data_length"], length_val)

    if "nullable" in col_map:
        _write_cell(ws, row, col_map["nullable"], "Y" if col.is_nullable else "N")
    elif "not_null" in col_map:
        _write_cell(ws, row, col_map["not_null"], "N" if col.is_nullable else "Y")

    if "pk" in col_map:
        _write_cell(ws, row, col_map["pk"], "Y" if col.is_pk else "N")
    if "key" in col_map:
        flags = []
        if col.is_pk:
            flags.append("PK")
        if col.is_fk:
            flags.append("FK")
        _write_cell(ws, row, col_map["key"], ", ".join(flags) if flags else None)

    if "fk" in col_map:
        _write_cell(
            ws, row, col_map["fk"], col.fk_ref or ("Y" if col.is_fk else "-")
        )

    if "index" in col_map:
        if col.index_key:
            _write_cell(ws, row, col_map["index"], col.index_key)
        else:
            _write_cell(ws, row, col_map["index"], "-")

    if "comment" in col_map:
        _write_cell(ws, row, col_map["comment"], col.extra_comment or None)


def _new_workbook(sheet_name: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _safe_sheet_title(sheet_name)
    ws.cell(1, 1).value = "테이블정의서"
    for col, label in enumerate(HEADER_LABELS, start=1):
        ws.cell(HEADER_ROW, col).value = label
    return wb


def _safe_sheet_title(name: str) -> str:
    text = (name or SHEET_NAME).strip() or SHEET_NAME
    return text[:31]


def inspect_template_layout(
    template: Path | BinaryIO,
    sheet_name: str | None = None,
) -> dict:
    """Return detected template format/columns for UI hints."""
    layout, _, _ = _read_template_layout(template, sheet_name)
    return {
        "format": layout.format,
        "sheet": layout.sheet_name,
        "has_db": layout.has_db,
        "has_schema": layout.has_schema,
    }
