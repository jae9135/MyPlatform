"""Query and upload table data (SELECT / INSERT only)."""

from __future__ import annotations

import csv
import io
from decimal import Decimal
from typing import Any
from uuid import UUID

import openpyxl
import psycopg2
from psycopg2.extras import RealDictCursor

from .db_client import get_connection_params
from .schema_reader import assert_ident, assert_user_schema, assert_writable_table

MAX_UPLOAD_BYTES = 2_000_000
MAX_UPLOAD_ROWS = 2_000
MAX_LIMIT = 200
DEFAULT_LIMIT = 100
CONFLICT_MODES = ("insert", "skip", "update", "renumber")


def query_table_data(
    schema: str,
    table: str,
    *,
    limit: int = DEFAULT_LIMIT,
    offset: int = 0,
    q: str = "",
) -> dict[str, Any]:
    schema_name = assert_user_schema(schema)
    table_name = assert_ident(table, "table")
    limit_n = max(1, min(int(limit or DEFAULT_LIMIT), MAX_LIMIT))
    offset_n = max(0, int(offset or 0))
    needle = (q or "").strip()

    conn = psycopg2.connect(**get_connection_params())
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute(
                """
                SELECT COUNT(*) AS n
                FROM information_schema.tables
                WHERE table_schema = %s
                  AND table_name = %s
                  AND table_type = 'BASE TABLE'
                """,
                (schema_name, table_name),
            )
            if int(cur.fetchone()["n"] or 0) == 0:
                raise ValueError(f"Table not found: {schema_name}.{table_name}")

            db_cols, pk_cols = _table_columns_and_pks(cur, schema_name, table_name)
            where_sql, where_params = _search_where(db_cols, needle)
            cur.execute(
                f'SELECT COUNT(*) AS n FROM "{schema_name}"."{table_name}"{where_sql}',
                where_params,
            )
            total = int(cur.fetchone()["n"] or 0)

            cur.execute(
                f'SELECT * FROM "{schema_name}"."{table_name}"{where_sql} '
                f"ORDER BY 1 LIMIT %s OFFSET %s",
                [*where_params, limit_n, offset_n],
            )
            raw_rows = cur.fetchall()
            columns = [d[0] for d in cur.description] if cur.description else []
            rows = [_serialize_dict(r, columns) for r in raw_rows]
            return {
                "schema": schema_name,
                "table": table_name,
                "columns": columns,
                "rows": rows,
                "total": total,
                "limit": limit_n,
                "offset": offset_n,
                "q": needle,
                "pk_columns": pk_cols,
            }
    finally:
        conn.close()


def upload_table_data(
    schema: str,
    table: str,
    file_bytes: bytes,
    filename: str = "",
    *,
    on_conflict: str = "skip",
) -> dict[str, Any]:
    schema_name, table_name = assert_writable_table(schema, table)
    mode = (on_conflict or "skip").strip().lower()
    if mode not in CONFLICT_MODES:
        raise ValueError(f"on_conflict must be one of {', '.join(CONFLICT_MODES)}")
    if not file_bytes:
        raise ValueError("Upload file is empty.")
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise ValueError(f"File too large (max {MAX_UPLOAD_BYTES} bytes).")

    rows = _parse_upload(file_bytes, filename)
    if not rows:
        raise ValueError("No data rows found in file.")
    if len(rows) > MAX_UPLOAD_ROWS:
        raise ValueError(f"Too many rows (max {MAX_UPLOAD_ROWS}).")

    conn = psycopg2.connect(**get_connection_params())
    inserted = 0
    updated = 0
    skipped = 0
    errors: list[str] = []
    try:
        with conn.cursor() as cur:
            db_cols, pk_cols = _table_columns_and_pks(cur, schema_name, table_name)
            mapped_rows, used_cols = _map_rows_to_columns(rows, db_cols)
            if not used_cols:
                raise ValueError("No matching columns between file and table.")
            if mode in ("skip", "update", "renumber") and not pk_cols:
                raise ValueError(
                    f"{mode} requires a primary key on {schema_name}.{table_name}."
                )
            if mode == "renumber" and len(pk_cols) != 1:
                raise ValueError("renumber requires a single-column primary key.")
            if mode == "renumber" and pk_cols[0] not in used_cols:
                used_cols = [pk_cols[0], *used_cols]

            next_pk = None
            if mode == "renumber":
                next_pk = _max_numeric_pk(cur, schema_name, table_name, pk_cols[0]) + 1

            for i, row in enumerate(mapped_rows, start=2):
                values = dict(row)
                if mode == "renumber":
                    values[pk_cols[0]] = str(next_pk)
                    next_pk += 1
                try:
                    status = _insert_row(
                        cur,
                        schema_name,
                        table_name,
                        used_cols,
                        values,
                        pk_cols,
                        mode,
                    )
                    if status == "inserted":
                        inserted += 1
                    elif status == "updated":
                        updated += 1
                    else:
                        skipped += 1
                except Exception as exc:
                    errors.append({"row": i, "message": str(exc)})
                    if len(errors) >= 50:
                        break
        if errors and inserted == 0 and updated == 0:
            conn.rollback()
            raise ValueError("; ".join(
                f"row {e['row']}: {e['message']}" for e in errors[:5]
            ))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return {
        "schema": schema_name,
        "table": table_name,
        "columns": used_cols,
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "error_csv": _errors_to_csv(errors) if errors else "",
        "on_conflict": mode,
    }


def _table_columns_and_pks(
    cur, schema: str, table: str
) -> tuple[list[str], list[str]]:
    cur.execute(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position
        """,
        (schema, table),
    )
    columns = []
    for r in cur.fetchall():
        columns.append(r["column_name"] if isinstance(r, dict) else r[0])
    if not columns:
        raise ValueError(f"Table not found: {schema}.{table}")
    cur.execute(
        """
        SELECT kcu.column_name
        FROM information_schema.table_constraints tc
        JOIN information_schema.key_column_usage kcu
          ON tc.constraint_name = kcu.constraint_name
         AND tc.table_schema = kcu.table_schema
        WHERE tc.constraint_type = 'PRIMARY KEY'
          AND tc.table_schema = %s
          AND tc.table_name = %s
        ORDER BY kcu.ordinal_position
        """,
        (schema, table),
    )
    pks = []
    for r in cur.fetchall():
        pks.append(r["column_name"] if isinstance(r, dict) else r[0])
    return columns, pks


def _map_rows_to_columns(
    rows: list[dict[str, Any]], db_cols: list[str]
) -> tuple[list[dict[str, Any]], list[str]]:
    lower_map = {c.lower(): c for c in db_cols}
    used: list[str] = []
    seen: set[str] = set()
    for key in rows[0].keys():
        actual = lower_map.get(str(key).strip().lower())
        if actual and actual not in seen:
            used.append(actual)
            seen.add(actual)
    mapped = []
    for row in rows:
        item: dict[str, Any] = {}
        lower_row = {str(k).strip().lower(): v for k, v in row.items()}
        for col in used:
            val = lower_row.get(col.lower())
            if val == "":
                val = None
            item[col] = val
        mapped.append(item)
    return mapped, used


def _insert_row(
    cur,
    schema: str,
    table: str,
    columns: list[str],
    values: dict[str, Any],
    pk_cols: list[str],
    mode: str,
) -> str:
    col_sql = ", ".join(f'"{c}"' for c in columns)
    placeholders = ", ".join(["%s"] * len(columns))
    params = [values.get(c) for c in columns]
    target = f'"{schema}"."{table}"'
    sql = f"INSERT INTO {target} ({col_sql}) VALUES ({placeholders})"

    if mode in ("insert", "renumber"):
        cur.execute(sql, params)
        return "inserted"

    pk_sql = ", ".join(f'"{c}"' for c in pk_cols)
    if mode == "skip":
        cur.execute(f"{sql} ON CONFLICT ({pk_sql}) DO NOTHING", params)
        return "inserted" if cur.rowcount else "skipped"

    non_pk = [c for c in columns if c not in pk_cols]
    if not non_pk:
        cur.execute(f"{sql} ON CONFLICT ({pk_sql}) DO NOTHING", params)
        return "inserted" if cur.rowcount else "skipped"
    set_sql = ", ".join(f'"{c}" = EXCLUDED."{c}"' for c in non_pk)
    cur.execute(
        f"{sql} ON CONFLICT ({pk_sql}) DO UPDATE SET {set_sql} "
        f"RETURNING (xmax = 0) AS is_insert",
        params,
    )
    flag = cur.fetchone()
    return "inserted" if flag and flag[0] else "updated"


def _max_numeric_pk(cur, schema: str, table: str, pk_col: str) -> int:
    assert_ident(pk_col, "column")
    cur.execute(
        f'SELECT COALESCE(MAX(({pk_col})::bigint), 0) '
        f'FROM "{schema}"."{table}" '
        f"WHERE ({pk_col})::text ~ '^[0-9]+$'"
    )
    row = cur.fetchone()
    return int(row[0] or 0) if row else 0


def _parse_upload(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    name = (filename or "").lower()
    if name.endswith(".csv") or _looks_like_csv(file_bytes):
        if name.endswith((".xlsx", ".xlsm")):
            return _parse_excel(file_bytes)
        try:
            return _parse_csv(file_bytes)
        except Exception:
            if name.endswith(".csv"):
                raise
            return _parse_excel(file_bytes)
    return _parse_excel(file_bytes)


def _looks_like_csv(file_bytes: bytes) -> bool:
    return not file_bytes.startswith(b"PK")


def _parse_csv(file_bytes: bytes) -> list[dict[str, Any]]:
    text = file_bytes.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    rows = []
    for row in reader:
        item = {}
        for k, v in row.items():
            if not k:
                continue
            item[str(k).strip()] = v.strip() if isinstance(v, str) else v
        if any(v not in (None, "") for v in item.values()):
            rows.append(item)
    return rows


def _parse_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    try:
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        headers = [h for h in headers if h]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None and v != "" for v in row):
                continue
            item = {}
            for i, header in enumerate(headers):
                val = row[i] if i < len(row) else None
                item[header] = val if val is not None else ""
            rows.append(item)
        return rows
    finally:
        wb.close()


def _serialize_dict(row: dict, columns: list[str]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for col in columns:
        out[col] = _serialize_value(row.get(col))
    return out


def _serialize_value(val: Any) -> Any:
    if val is None:
        return None
    if isinstance(val, Decimal):
        if val == val.to_integral_value():
            return int(val)
        return float(val)
    if isinstance(val, UUID):
        return str(val)
    if isinstance(val, (bytes, memoryview)):
        return bytes(val).hex()
    if hasattr(val, "isoformat"):
        try:
            return val.isoformat(sep=" ", timespec="seconds")
        except TypeError:
            return val.isoformat()
    if isinstance(val, (str, int, float, bool)):
        return val
    return str(val)


def preview_upload(
    schema: str,
    table: str,
    file_bytes: bytes,
    filename: str = "",
    *,
    preview_rows: int = 20,
) -> dict[str, Any]:
    schema_name, table_name = assert_writable_table(schema, table)
    if not file_bytes:
        raise ValueError("Upload file is empty.")
    rows = _parse_upload(file_bytes, filename)
    if not rows:
        raise ValueError("No data rows found in file.")
    conn = psycopg2.connect(**get_connection_params())
    try:
        with conn.cursor() as cur:
            db_cols, pk_cols = _table_columns_and_pks(cur, schema_name, table_name)
    finally:
        conn.close()
    mapped, used_cols = _map_rows_to_columns(rows, db_cols)
    skipped_headers = [
        str(k).strip()
        for k in rows[0].keys()
        if str(k).strip() and str(k).strip().lower() not in {c.lower() for c in used_cols}
    ]
    sample = [
        {c: _serialize_value(r.get(c)) for c in used_cols}
        for r in mapped[:preview_rows]
    ]
    return {
        "schema": schema_name,
        "table": table_name,
        "columns": used_cols,
        "pk_columns": pk_cols,
        "skipped_headers": skipped_headers,
        "row_count": len(mapped),
        "preview": sample,
    }


def update_table_row(
    schema: str,
    table: str,
    pk: dict[str, Any],
    values: dict[str, Any],
) -> dict[str, Any]:
    schema_name, table_name = assert_writable_table(schema, table)
    conn = psycopg2.connect(**get_connection_params())
    try:
        with conn.cursor() as cur:
            db_cols, pk_cols = _table_columns_and_pks(cur, schema_name, table_name)
            if not pk_cols:
                raise ValueError("Table has no primary key; cannot update a single row.")
            pk_vals = _require_pk(pk, pk_cols)
            assigned = []
            params: list[Any] = []
            lower_vals = {str(k).lower(): v for k, v in (values or {}).items()}
            for col in db_cols:
                if col in pk_cols:
                    continue
                if col.lower() not in lower_vals:
                    continue
                assigned.append(f'"{col}" = %s')
                params.append(lower_vals[col.lower()])
            if not assigned:
                raise ValueError("No updatable columns in payload.")
            where, where_params = _pk_where(pk_cols, pk_vals)
            cur.execute(
                f'UPDATE "{schema_name}"."{table_name}" SET {", ".join(assigned)} {where}',
                [*params, *where_params],
            )
            if cur.rowcount == 0:
                raise ValueError("Row not found for the given primary key.")
        conn.commit()
        return {"updated": 1, "pk": pk_vals}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def delete_table_row(schema: str, table: str, pk: dict[str, Any]) -> dict[str, Any]:
    schema_name, table_name = assert_writable_table(schema, table)
    conn = psycopg2.connect(**get_connection_params())
    try:
        with conn.cursor() as cur:
            _cols, pk_cols = _table_columns_and_pks(cur, schema_name, table_name)
            if not pk_cols:
                raise ValueError("Table has no primary key; cannot delete a single row.")
            pk_vals = _require_pk(pk, pk_cols)
            where, where_params = _pk_where(pk_cols, pk_vals)
            cur.execute(
                f'DELETE FROM "{schema_name}"."{table_name}" {where}',
                where_params,
            )
            if cur.rowcount == 0:
                raise ValueError("Row not found for the given primary key.")
        conn.commit()
        return {"deleted": 1, "pk": pk_vals}
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def export_table_data(
    schema: str,
    table: str,
    *,
    q: str = "",
    fmt: str = "csv",
    max_rows: int = 5_000,
) -> tuple[bytes, str, str]:
    data = query_table_data(schema, table, limit=min(max_rows, 200), offset=0, q=q)
    rows = list(data["rows"])
    total = min(int(data["total"]), max_rows)
    cols = data["columns"]
    offset = data["limit"]
    while offset < total:
        chunk = query_table_data(
            schema, table, limit=min(200, total - offset), offset=offset, q=q
        )
        if not chunk["rows"]:
            break
        rows.extend(chunk["rows"])
        offset += len(chunk["rows"])
    name = f"{schema}_{table}"
    if fmt == "xlsx":
        return (
            _rows_to_xlsx(cols, rows),
            f"{name}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    return _rows_to_csv(cols, rows), f"{name}.csv", "text/csv; charset=utf-8"


def _search_where(columns: list[str], needle: str) -> tuple[str, list[Any]]:
    if not needle:
        return "", []
    like = f"%{needle}%"
    parts = [f'CAST("{c}" AS TEXT) ILIKE %s' for c in columns]
    return " WHERE (" + " OR ".join(parts) + ")", [like] * len(columns)


def _require_pk(pk: dict[str, Any], pk_cols: list[str]) -> dict[str, Any]:
    lower = {str(k).lower(): v for k, v in (pk or {}).items()}
    out = {}
    for col in pk_cols:
        if col.lower() not in lower:
            raise ValueError(f"Missing primary key value: {col}")
        out[col] = lower[col.lower()]
    return out


def _pk_where(pk_cols: list[str], pk_vals: dict[str, Any]) -> tuple[str, list[Any]]:
    clauses = [f'"{c}" = %s' for c in pk_cols]
    params = [pk_vals[c] for c in pk_cols]
    return " WHERE " + " AND ".join(clauses), params


def _errors_to_csv(errors: list[dict[str, Any]]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["row", "message"])
    for err in errors:
        writer.writerow([err.get("row"), err.get("message")])
    return buf.getvalue()


def _rows_to_csv(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(["" if row.get(c) is None else row.get(c) for c in columns])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def _rows_to_xlsx(columns: list[str], rows: list[dict[str, Any]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(columns)
    for row in rows:
        ws.append([row.get(c) for c in columns])
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()
