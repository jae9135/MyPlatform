"""Parse table definitions from Excel workbook (multiple template variants)."""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import openpyxl

# Legacy fixed layout (excel_writer still uses these indices)
COL_DB = 2
COL_SCHEMA = 3
COL_TABLE_KO = 4
COL_TABLE_EN = 5
COL_COLUMN_KO = 6
COL_COLUMN_EN = 7
COL_DATA_TYPE = 8
COL_DATA_LENGTH = 9
COL_NOT_NULL = 10
COL_PK = 11
COL_FK = 12
COL_INDEX = 14
COL_COMMENT = 16
HEADER_ROW = 3
DATA_START_ROW = 4

DEFAULT_SHEET = "테이블정의서"
SHEET_HINTS = (
    "테이블정의서",
    "테이블 정의서",
    "테이블명세서",
    "테이블 목록",
)

INSTRUCTION_MARKERS = ("[작성", "작성 방법", "작성 사례", "입력 (*", "입력\n", "아래의")

_CONSTRAINT_LINE = re.compile(r"^(PK|FK|UK|IX)_([^(]+)\(([^)]*)\)$", re.I)

LABEL_TO_FIELD: dict[str, str] = {
    "no": "no",
    "번호": "no",
    "db명": "db",
    "스키마명": "schema",
    "한글테이블명": "table_ko",
    "영문테이블명": "table_en",
    "한글컬럼명": "column_ko",
    "영문컬럼명": "column_en",
    "필드id": "column_en",
    "필드명": "column_ko",
    "데이터타입": "data_type",
    "type": "data_type",
    "데이터길이": "data_length",
    "길이": "data_length",
    "length": "data_length",
    "datalength": "data_length",
    "notnull여부": "not_null",
    "null여부": "nullable",
    "pk여부": "pk",
    "key": "key",
    "fk여부": "fk",
    "indexkey": "index",
    "코멘트": "comment",
    "비고": "comment",
    "디폴트값": "default",
    "기본값": "default",
}


@dataclass
class ColumnDef:
    name: str
    korean_name: str
    data_type: str
    length: int | None
    not_null: bool
    is_pk: bool
    comment: str | None = None
    is_fk: bool = False
    fk_ref: str | None = None
    index_key: str | None = None
    is_uk: bool = False
    default_value: str | None = None


@dataclass
class TableDef:
    db_name: str
    schema: str
    name: str
    korean_name: str
    columns: list[ColumnDef] = field(default_factory=list)

    @property
    def pk_columns(self) -> list[str]:
        return [col.name for col in self.columns if col.is_pk]


@dataclass
class ParseMeta:
    sheet_name: str
    format: str  # flat | block
    tables: list[TableDef]
    system_name: str = ""
    created_date: str = ""
    author: str = ""


def parse_excel(
    source: Path | BinaryIO, sheet_name: str | None = DEFAULT_SHEET
) -> list[TableDef]:
    """Read table/column definitions from Excel and return grouped TableDef list."""
    return parse_excel_with_meta(source, sheet_name).tables


def parse_excel_with_meta(
    source: Path | BinaryIO, sheet_name: str | None = DEFAULT_SHEET
) -> ParseMeta:
    wb = openpyxl.load_workbook(source, data_only=True)
    resolved = resolve_sheet(wb, sheet_name)
    ws = wb[resolved]
    tables, fmt = _parse_worksheet(ws)
    system_name = _read_system_name(ws, fmt)
    created_date, author = _read_doc_meta(ws, fmt)
    wb.close()
    if not tables:
        raise ValueError(
            f"시트 '{resolved}'에서 테이블/컬럼 정의를 찾지 못했습니다. "
            "지원 양식: 목록형(테이블정의서) 또는 블록형(테이블 정의서/테이블명세서)."
        )
    return ParseMeta(
        sheet_name=resolved,
        format=fmt,
        tables=tables,
        system_name=system_name,
        created_date=created_date,
        author=author,
    )


def resolve_sheet(wb, sheet_name: str | None) -> str:
    if sheet_name and str(sheet_name).strip():
        requested = str(sheet_name).strip()
        if requested in wb.sheetnames:
            return requested
        req_norm = normalize_label(requested)
        for name in wb.sheetnames:
            if normalize_label(name) == req_norm:
                return name

    best_name = ""
    best_score = 0
    for name in wb.sheetnames:
        score = _score_design_sheet(wb[name])
        if score > best_score:
            best_score = score
            best_name = name

    if best_score >= 4 and best_name:
        return best_name

    available = ", ".join(wb.sheetnames)
    if sheet_name and str(sheet_name).strip():
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {available}"
        )
    raise ValueError(
        f"테이블정의서 시트를 찾지 못했습니다. Available: {available}"
    )


def table_defs_to_dataframe_rows(tables: list[TableDef]) -> list[dict]:
    rows: list[dict] = []
    no = 1
    for table in tables:
        for col in table.columns:
            rows.append(
                {
                    "No": no,
                    "DB명": table.db_name.upper(),
                    "스키마명": table.schema,
                    "한글 테이블명": table.korean_name,
                    "영문 테이블명": table.name.upper(),
                    "한글 컬럼명": col.korean_name,
                    "영문 컬럼명": col.name.upper(),
                    "데이터 타입": col.data_type,
                    "데이터 길이": col.length,
                    "Not Null 여부": "Y" if col.not_null else "N",
                    "PK 여부": "Y" if col.is_pk else "N",
                }
            )
            no += 1
    return rows


def _parse_worksheet(ws) -> tuple[list[TableDef], str]:
    flat = _find_flat_header(ws)
    if flat:
        header_row, col_map = flat
        tables = _parse_flat(ws, header_row, col_map)
        if tables:
            return tables, "flat"

    tables = _parse_block(ws)
    if tables:
        return tables, "block"
    return [], "unknown"


def _find_flat_header(ws) -> tuple[int, dict[str, int]] | None:
    max_scan = min(25, ws.max_row or 0)
    for row in range(1, max_scan + 1):
        col_map = _map_header_row(ws, row)
        if {"table_en", "column_en"}.issubset(col_map.keys()):
            return row, col_map
    return None


def _parse_flat(ws, header_row: int, col_map: dict[str, int]) -> list[TableDef]:
    tables: dict[str, TableDef] = {}
    default_col = col_map.get("default")

    for row in range(header_row + 1, (ws.max_row or 0) + 1):
        table_name = _cell(ws, row, col_map.get("table_en", COL_TABLE_EN))
        column_name = _cell(ws, row, col_map.get("column_en", COL_COLUMN_EN))
        data_type = _cell(ws, row, col_map.get("data_type", COL_DATA_TYPE))

        if not table_name or not column_name:
            continue
        if _is_instruction_text(table_name) or _is_instruction_text(column_name):
            continue
        if normalize_label(data_type) in ("데이터타입", "type"):
            continue

        table_key = str(table_name).strip()
        if table_key not in tables:
            tables[table_key] = TableDef(
                db_name="dbm",
                schema="db1",
                name=table_key.lower(),
                korean_name=str(
                    _cell(ws, row, col_map.get("table_ko", COL_TABLE_KO)) or table_key
                ).strip(),
            )

        not_null_col = col_map.get("not_null")
        nullable_col = col_map.get("nullable")
        pk_col = col_map.get("pk")
        key_col = col_map.get("key")
        fk_col = col_map.get("fk")
        index_col = col_map.get("index", COL_INDEX)
        comment_col = col_map.get("comment", COL_COMMENT)

        not_null = False
        if nullable_col:
            not_null = _parse_not_null("nullable", _cell(ws, row, nullable_col))
        elif not_null_col:
            not_null = _parse_not_null("not_null", _cell(ws, row, not_null_col))

        pk_val = _cell(ws, row, pk_col) if pk_col else None
        key_val = _cell(ws, row, key_col) if key_col else None
        fk_val = _cell(ws, row, fk_col) if fk_col else None
        is_pk, is_fk = _parse_key_flags(key_val, pk_val, fk_val)
        dtype, dlen = _split_excel_type(
            data_type,
            _parse_length(_effective_cell(ws, row, col_map.get("data_length", COL_DATA_LENGTH))),
        )

        tables[table_key].columns.append(
            ColumnDef(
                name=str(column_name).strip().lower(),
                korean_name=str(
                    _cell(ws, row, col_map.get("column_ko", COL_COLUMN_KO)) or column_name
                ).strip(),
                data_type=dtype,
                length=dlen,
                not_null=not_null or is_pk,
                is_pk=is_pk,
                comment=_optional_str(_cell(ws, row, comment_col)),
                is_fk=is_fk,
                fk_ref=_fk_ref(fk_val),
                index_key=_index_key(_cell(ws, row, index_col)),
                is_uk=_is_uk(_cell(ws, row, index_col)),
                default_value=_optional_str(_cell(ws, row, default_col))
                if default_col
                else None,
            )
        )

    return list(tables.values())


def _parse_block(ws) -> list[TableDef]:
    tables: dict[str, TableDef] = {}
    row = 1
    max_row = ws.max_row or 0

    while row <= max_row:
        if not _is_table_meta_row(ws, row):
            row += 1
            continue

        table_ko = _cell(ws, row, 3)
        table_en = _cell(ws, row, 6)
        if _is_instruction_text(table_ko) or _is_instruction_text(table_en):
            row += 1
            continue

        _read_module_name(ws, row)
        header_row = row + 1
        if header_row > max_row or not _is_block_column_header(ws, header_row):
            row += 1
            continue

        col_map = _map_header_row(ws, header_row)
        data_row = header_row + 1
        table_key = str(table_en).strip()
        if table_key not in tables:
            tables[table_key] = TableDef(
                db_name="dbm",
                schema="db1",
                name=table_key.lower(),
                korean_name=str(table_ko or table_key).strip(),
            )

        index_key_parts: list[str] = []
        while data_row <= max_row:
            marker = str(_effective_cell(ws, data_row, 1) or "").strip()
            if _is_table_meta_row(ws, data_row):
                break
            if marker == "Index Key" or (
                not marker
                and _looks_like_constraint(_read_block_index_key_value(ws, data_row))
            ):
                val = _read_block_index_key_value(ws, data_row)
                if val:
                    index_key_parts.append(val)
                data_row += 1
                continue
            if marker in ("업무규칙", "테이블 정의서", "테이블정의서"):
                data_row += 1
                break

            column_name = _effective_cell(ws, data_row, col_map.get("column_en", 2))
            if not column_name or _is_instruction_text(column_name):
                data_row += 1
                continue

            column_ko = _effective_cell(ws, data_row, col_map.get("column_ko", 4))
            data_type = _effective_cell(ws, data_row, col_map.get("data_type", 5))
            if not data_type or _is_instruction_text(data_type):
                data_row += 1
                continue

            nullable_col = col_map.get("nullable")
            not_null_col = col_map.get("not_null")
            key_col = col_map.get("key")
            pk_col = col_map.get("pk")
            fk_col = col_map.get("fk")
            comment_col = col_map.get("comment")
            default_col = col_map.get("default")

            not_null = False
            if nullable_col:
                not_null = _parse_not_null("nullable", _cell(ws, data_row, nullable_col))
            elif not_null_col:
                not_null = _parse_not_null(
                    "not_null", _cell(ws, data_row, not_null_col)
                )

            key_val = _effective_cell(ws, data_row, key_col) if key_col else None
            pk_val = _effective_cell(ws, data_row, pk_col) if pk_col else None
            fk_val = _effective_cell(ws, data_row, fk_col) if fk_col else None
            is_pk, is_fk = _parse_key_flags(key_val, pk_val, fk_val)
            dtype, dlen = _split_excel_type(
                data_type,
                _parse_length(
                    _effective_cell(ws, data_row, col_map.get("data_length", 6))
                ),
            )

            tables[table_key].columns.append(
                ColumnDef(
                    name=str(column_name).strip().lower(),
                    korean_name=str(column_ko or column_name).strip(),
                    data_type=dtype,
                    length=dlen,
                    not_null=not_null or is_pk,
                    is_pk=is_pk,
                    comment=_optional_str(
                        _cell(ws, data_row, comment_col) if comment_col else None
                    ),
                    is_fk=is_fk,
                    fk_ref=_fk_ref(fk_val),
                    index_key=None,
                    is_uk=False,
                    default_value=_optional_str(
                        _cell(ws, data_row, default_col) if default_col else None
                    ),
                )
            )
            data_row += 1

        if index_key_parts:
            _apply_index_key_constraints(
                tables[table_key], "\n".join(index_key_parts), tables
            )
        row = data_row

    _infer_missing_fk_refs(tables)
    return list(tables.values())


def _score_design_sheet(ws) -> int:
    score = 0
    max_scan = min(30, ws.max_row or 0)
    for row in range(1, max_scan + 1):
        col_map = _map_header_row(ws, row)
        if {"table_en", "column_en"}.issubset(col_map.keys()):
            score += 6
        if "column_ko" in col_map and "data_type" in col_map:
            score += 2
        if _is_block_column_header(ws, row):
            score += 5
        if _is_table_meta_row(ws, row):
            score += 1
    name = normalize_label(ws.title)
    if "테이블정의서" in name or name in ("테이블정의서", "테이블명세서"):
        score += 3
    return score


def _map_header_row(ws, row: int) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        label = normalize_label(_effective_cell(ws, row, col))
        if not label:
            continue
        field = LABEL_TO_FIELD.get(label)
        if field and field not in mapping:
            mapping[field] = col
    return mapping


def _is_table_meta_row(ws, row: int) -> bool:
    c1 = normalize_label(_cell(ws, row, 1))
    c5 = normalize_label(_cell(ws, row, 5))
    return c1 == "테이블명" and c5 == "테이블id"


def _is_block_column_header(ws, row: int) -> bool:
    col_map = _map_header_row(ws, row)
    return "column_en" in col_map and "column_ko" in col_map and "data_type" in col_map


def _read_system_name(ws, fmt: str) -> str:
    return _read_labeled_meta(ws, fmt, {"시스템명", "모듈시스템명"})


def _read_doc_meta(ws, fmt: str) -> tuple[str, str]:
    created = _read_labeled_meta(ws, fmt, {"작성일", "작성일자"})
    author = _read_labeled_meta(ws, fmt, {"작성자", "작성자명"})
    return _format_meta_date(created), author


def _format_meta_date(value: str) -> str:
    text = (value or "").strip()
    if not text:
        return ""
    match = re.match(r"^(\d{4})[./-](\d{1,2})[./-](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return text


def _read_labeled_meta(ws, fmt: str, labels: set[str]) -> str:
    max_col = ws.max_column or 1
    if fmt == "block":
        last = min(8, ws.max_row or 1)
        for row in range(1, last + 1):
            if _is_table_meta_row(ws, row):
                last = max(1, row - 1)
                break
    else:
        header = _find_flat_header(ws)
        last = header[0] if header else min(12, ws.max_row or 1)
    for row in range(1, last + 1):
        for col in range(1, max_col + 1):
            if normalize_label(_effective_cell(ws, row, col)) not in labels:
                continue
            for c in range(col + 1, min(col + 8, max_col + 1)):
                val = _effective_cell(ws, row, c)
                if val in (None, "") or _is_instruction_text(val):
                    continue
                if normalize_label(val) in {
                    "시스템명",
                    "모듈시스템명",
                    "작성일",
                    "작성일자",
                    "작성자",
                    "작성자명",
                }:
                    continue
                if hasattr(val, "strftime"):
                    return val.strftime("%Y-%m-%d")
                return str(val).strip()
    if fmt == "block":
        max_row = ws.max_row or 0
        for row in range(1, max_row + 1):
            if not _is_table_meta_row(ws, row):
                continue
            name = _read_module_labeled(ws, row, labels)
            if name:
                return name
    return ""


def _read_module_labeled(ws, table_meta_row: int, labels: set[str]) -> str:
    module_row = table_meta_row - 1
    if module_row < 1:
        return ""
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        if normalize_label(_effective_cell(ws, module_row, col)) not in labels:
            continue
        for c in range(col + 1, min(col + 8, max_col + 1)):
            val = _effective_cell(ws, module_row, c)
            if val in (None, "") or _is_instruction_text(val):
                continue
            if normalize_label(val) in labels:
                continue
            if hasattr(val, "strftime"):
                return val.strftime("%Y-%m-%d")
            return str(val).strip()
    return ""


def _read_module_name(ws, table_meta_row: int) -> str | None:
    module_row = table_meta_row - 1
    if module_row < 1:
        return None
    if normalize_label(_cell(ws, module_row, 1)) != "모듈시스템명":
        return None
    value = _cell(ws, module_row, 3)
    if _is_instruction_text(value):
        return None
    return str(value).strip()


def normalize_label(value) -> str:
    text = str(value or "").replace("\n", " ").strip().lower()
    text = re.sub(r"\s+", "", text)
    text = text.replace("(", "").replace(")", "")
    return text


def _is_instruction_text(value) -> bool:
    text = str(value or "").strip()
    if not text:
        return True
    if len(text) > 100:
        return True
    return any(marker in text for marker in INSTRUCTION_MARKERS)


def _cell(ws, row: int, col: int | None):
    if not col:
        return None
    return ws.cell(row, col).value


def _effective_cell(ws, row: int, col: int | None):
    if not col:
        return None
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col).value
    return ws.cell(row, col).value


def _looks_like_constraint(value) -> bool:
    text = str(value or "").strip()
    return bool(re.match(r"^(PK|FK|UK|IX)_", text, re.I))


def _normalize_identifier(value: str) -> str:
    return str(value).strip().lower()


def _parse_length(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return int(value)
    text = str(value).strip().replace(",", "")
    if not text or text in ("-", "/", "None", "NONE"):
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    try:
        return int(match.group(1))
    except (TypeError, ValueError):
        return None


def _split_excel_type(data_type, length: int | None) -> tuple[str, int | None]:
    text = str(data_type or "").strip()
    packed = re.match(r"^([A-Za-z][A-Za-z0-9_]+)\s*\(\s*(\d+)\s*\)\s*$", text)
    if packed:
        name = packed.group(1).upper()
        packed_len = int(packed.group(2))
        return name, length if length else packed_len
    return text, length


def _parse_not_null(kind: str, value) -> bool:
    text = str(value or "").strip().upper()
    if not text or text in ("-", "NONE", "NULL"):
        return False
    if kind == "nullable":
        return text in ("N", "NO")
    return text == "Y"


def _parse_pk(key_val, pk_val) -> bool:
    is_pk, _ = _parse_key_flags(key_val, pk_val, None)
    return is_pk


def _parse_key_flags(key_val, pk_val, fk_val) -> tuple[bool, bool]:
    text = str(key_val or "").upper()
    tokens = set(re.findall(r"PK|FK|UK", text))
    is_pk = "PK" in tokens or _is_yes(pk_val)
    is_fk = "FK" in tokens or _is_fk(fk_val)
    return is_pk, is_fk


def _read_block_index_key_value(ws, row: int) -> str | None:
    value_col = 3
    for col in range(2, (ws.max_column or 0) + 1):
        if _cell(ws, row, col) not in (None, ""):
            value_col = col
            break
    return _optional_str(_cell(ws, row, value_col))


def _split_index_key_text(text: str) -> list[str]:
    lines: list[str] = []
    for line in str(text or "").replace("\r\n", "\n").split("\n"):
        for part in line.split(";"):
            chunk = part.strip()
            if chunk:
                lines.append(chunk)
    return lines


def _parse_constraint_line(line: str) -> tuple[str, str, list[str]] | None:
    match = _CONSTRAINT_LINE.match(line.strip())
    if not match:
        return None
    kind = match.group(1).upper()
    name = match.group(2).strip()
    cols = [c.strip().lower() for c in match.group(3).split(",") if c.strip()]
    return kind, name, cols


def _resolve_column_name(table: TableDef, raw: str) -> str | None:
    want = raw.strip().lower()
    names = [c.name for c in table.columns]
    if want in names:
        return want
    for name in names:
        if name.replace("_", "") == want.replace("_", ""):
            return name
    for name in names:
        if want in name or name in want:
            return name
    return None


def _find_table_by_token(token: str, tables: dict[str, TableDef]) -> TableDef | None:
    needle = token.strip().lower()
    for key, table in tables.items():
        if key.lower() == needle or table.name == needle:
            return table
    return None


def _infer_fk_parent_table(
    constraint_name: str, child_table: TableDef, tables: dict[str, TableDef]
) -> TableDef | None:
    body = constraint_name.strip()
    if body.upper().startswith("FK_"):
        body = body[3:]
    child_up = child_table.name.upper().replace("-", "_")
    body_up = body.upper()
    if body_up.startswith(child_up + "_"):
        parent_token = body[len(child_up) + 1 :]
        parent = _find_table_by_token(parent_token.split("_")[0], tables)
        if parent:
            return parent
        parent = _find_table_by_token(parent_token, tables)
        if parent:
            return parent
    if "_" in body:
        parent = _find_table_by_token(body.split("_")[-1], tables)
        if parent:
            return parent
    return None


def _build_fk_ref(parent_table: TableDef, fk_col: str) -> str:
    parent_cols = {c.name for c in parent_table.columns}
    if fk_col in parent_cols:
        return f"{parent_table.name}({fk_col})"
    pk_cols = parent_table.pk_columns
    if pk_cols:
        return f"{parent_table.name}({pk_cols[0]})"
    return f"{parent_table.name}({fk_col})"


def _apply_index_key_constraints(
    table: TableDef, text: str, tables: dict[str, TableDef]
) -> None:
    by_name = {c.name: c for c in table.columns}
    for line in _split_index_key_text(text):
        parsed = _parse_constraint_line(line)
        if not parsed:
            continue
        kind, name, raw_cols = parsed
        resolved_cols = [_resolve_column_name(table, col) for col in raw_cols]
        resolved_cols = [col for col in resolved_cols if col]
        if kind == "PK":
            for col_name in resolved_cols:
                col = by_name.get(col_name)
                if not col:
                    continue
                col.is_pk = True
                col.not_null = True
                if not col.index_key:
                    col.index_key = line
        elif kind == "FK":
            parent = _infer_fk_parent_table(name, table, tables)
            target_cols = resolved_cols or list(by_name.keys())
            for col_name in target_cols:
                col = by_name.get(col_name)
                if not col:
                    continue
                col.is_fk = True
                if parent:
                    col.fk_ref = _build_fk_ref(parent, col_name)
                if col.index_key and col.index_key != line:
                    col.index_key = f"{col.index_key}\n{line}"
                else:
                    col.index_key = line
        elif kind == "UK":
            for col_name in resolved_cols:
                col = by_name.get(col_name)
                if not col:
                    continue
                col.is_uk = True
                col.index_key = line


def _infer_missing_fk_refs(tables: dict[str, TableDef]) -> None:
    for table in tables.values():
        for col in table.columns:
            if not col.is_fk or col.fk_ref:
                continue
            if not col.name.endswith("_id"):
                continue
            candidate = col.name[:-3]
            parent = _find_table_by_token(candidate, tables)
            if not parent:
                continue
            col.fk_ref = _build_fk_ref(parent, col.name)


def _is_yes(value) -> bool:
    return str(value or "").strip().upper() == "Y"


def _optional_str(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _blankish(value) -> bool:
    text = str(value or "").strip().upper()
    return text in ("", "-", "N", "NONE", "NULL")


def _is_fk(value) -> bool:
    return not _blankish(value)


def _fk_ref(value) -> str | None:
    if _blankish(value):
        return None
    text = str(value).strip()
    if text.upper() in ("Y", "YES", "FK"):
        return None
    return text


def _index_key(value) -> str | None:
    if _blankish(value):
        return None
    text = str(value).strip()
    if text.upper().startswith("PK_"):
        return text
    return text


def _is_uk(value) -> bool:
    text = str(value or "").strip().upper()
    return text.startswith("UK") or "UNIQUE" in text


def _find_header_col(ws, names: tuple[str, ...]) -> int | None:
    lowered = tuple(n.lower() for n in names)
    for col in range(1, 24):
        label = str(ws.cell(HEADER_ROW, col).value or "").strip()
        if not label:
            continue
        if any(n in label.lower() for n in lowered):
            return col
    return None
