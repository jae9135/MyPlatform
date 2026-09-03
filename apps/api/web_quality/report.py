from __future__ import annotations

import base64
import html
import io
import re
import zipfile
from typing import Any, Literal

import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def _screenshot_map(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {s["id"]: s for s in (payload.get("screenshots") or []) if s.get("id")}


def _screen_index_map(payload: dict[str, Any]) -> dict[str, int]:
    order: list[str] = []
    coverage = payload.get("coverage") or {}
    for s in coverage.get("screens") or []:
        sid = s.get("state_id")
        if sid and sid not in order:
            order.append(str(sid))
    for shot in payload.get("screenshots") or []:
        if shot.get("kind") == "state":
            sid = shot.get("state_id")
            if sid and sid not in order:
                order.append(str(sid))
    return {sid: i + 1 for i, sid in enumerate(order)}


def _screen_no_label(screen_map: dict[str, int], state_id: Any) -> str:
    idx = screen_map.get(str(state_id or ""), 0)
    return f"#{idx}" if idx else ""


def _resolve_finding_state_id(finding: dict[str, Any]) -> str:
    if finding.get("state_id"):
        return str(finding["state_id"])
    target = str(finding.get("target") or "")
    location = str(finding.get("location") or "")
    if target == "screen" and location and "::" not in location:
        return location
    m = re.search(r"::\s*([a-z0-9_]+)\s*::", location, re.I)
    if m:
        return m.group(1)
    for part in location.split("::"):
        p = part.strip()
        if re.match(r"^[a-z][a-z0-9_]*$", p, re.I) and not p.lower().startswith("http"):
            return p
    return ""


def _finding_screen_no_label(
    finding: dict[str, Any], screen_map: dict[str, int]
) -> str:
    return _screen_no_label(screen_map, _resolve_finding_state_id(finding))


def _finding_screen_no_html(
    finding: dict[str, Any], screen_map: dict[str, int]
) -> str:
    label = _finding_screen_no_label(finding, screen_map)
    if label:
        return f'<span class="screen-no">{html.escape(label)}</span>'
    return "—"


def _image_src(shot: dict[str, Any], mode: Literal["embed", "relative"]) -> str:
    if mode == "relative":
        return html.escape(str(shot.get("filename", "")))
    return html.escape(str(shot.get("data_url", "")))


def _enriched_finding(finding: dict[str, Any]) -> dict[str, Any]:
    from web_quality.fix_guides import enrich_finding

    return enrich_finding(finding)


def _ref_link_items_from(f: dict[str, Any]) -> list[tuple[str, str]]:
    from web_quality.ref_links import GUIDELINE_UIUX_2025_URL

    items: list[tuple[str, str]] = []
    primary = str(f.get("ref_url") or "").strip()
    fallback = str(f.get("ref_fallback_url") or "").strip()
    rule_id = str(f.get("rule_id") or "")
    category = str(f.get("category") or "")

    if rule_id.startswith("UX-KRDS-"):
        primary_label = "KRDS 관련근거"
    elif category == "a11y" or f.get("kwcag_id") or re.match(r"^\d+\.\d+", rule_id):
        primary_label = "KWCAG 관련근거"
    else:
        primary_label = "관련근거"

    if primary:
        items.append((primary_label, primary))
    primary_base = primary.split("#", 1)[0] if primary else ""
    fallback_base = fallback.split("#", 1)[0] if fallback else ""
    if fallback and fallback_base != primary_base:
        is_a11y = (
            category == "a11y"
            or f.get("kwcag_id")
            or rule_id.startswith("WA-")
            or re.match(r"^\d+\.\d+", rule_id)
        )
        is_krds = rule_id.startswith("UX-KRDS-") or category == "uiux"
        if is_a11y or "kwcag" in fallback.lower() or "a11ykr.github.io" in fallback.lower():
            fb_label = "KWCAG 목록 (대체)"
        elif is_krds:
            fb_label = "KRDS 목록 (대체)"
        else:
            fb_label = "목록 (대체)"
        items.append((fb_label, fallback))
    guideline = str(f.get("guideline_url") or "").strip()
    if not guideline and (rule_id.startswith("UX-KRDS-") or category == "uiux"):
        guideline = GUIDELINE_UIUX_2025_URL
    if guideline and guideline not in {primary, fallback}:
        items.append(("UI·UX 가이드라인(2025.08)", guideline))
    return items


def _ref_link_items(finding: dict[str, Any]) -> list[tuple[str, str]]:
    return _ref_link_items_from(_enriched_finding(finding))


def _ref_text(finding: dict[str, Any]) -> str:
    f = _enriched_finding(finding)
    krds_ref = str(f.get("krds_ref") or "").strip()
    items = _ref_link_items_from(f)
    if krds_ref and items:
        return f"{krds_ref}\n" + "\n".join(f"{label}: {url}" for label, url in items[1:])
    if krds_ref:
        return krds_ref
    if items:
        return "\n".join(f"{label}: {url}" for label, url in items)
    return ""


def _ref_cell_html(finding: dict[str, Any]) -> str:
    f = _enriched_finding(finding)
    items = _ref_link_items_from(f)
    krds_ref = str(f.get("krds_ref") or "").strip()
    if not items and not krds_ref:
        return "—"

    parts: list[str] = []
    primary_url = items[0][1] if items else ""
    extra = items[1:] if krds_ref and primary_url and items else items

    if krds_ref:
        if primary_url:
            parts.append(
                f'<a class="ref-link ref-krds" href="{html.escape(primary_url)}" '
                f'target="_blank" rel="noopener">{html.escape(krds_ref)}</a>'
            )
        else:
            parts.append(f'<span class="muted">{html.escape(krds_ref)}</span>')
    elif items:
        label, url = items[0]
        parts.append(
            f'<a class="ref-link" href="{html.escape(url)}" '
            f'target="_blank" rel="noopener">{html.escape(label)}</a>'
        )
        extra = items[1:]

    for label, url in extra:
        css = "ref-link ref-fallback" if "대체" in label else "ref-link"
        parts.append(
            f'<div class="ref-extra"><a class="{css}" href="{html.escape(url)}" '
            f'target="_blank" rel="noopener">{html.escape(label)}</a></div>'
        )
    return "".join(parts) if parts else "—"


def _findings_df(
    findings: list[dict[str, Any]], screen_map: dict[str, int] | None = None
) -> pd.DataFrame:
    smap = screen_map or {}
    rows = []
    for f in findings:
        fix = f.get("fix", "")
        ref = f.get("fix_url", "")
        if ref and ref not in str(fix):
            fix = f"{fix}\n참고: {ref}".strip() if fix else str(ref)
        rows.append(
            {
                "대상유형": "소스" if f.get("target") == "source" else "화면",
                "화면번호": _finding_screen_no_label(f, smap) if smap else "",
                "위치": f.get("location", ""),
                "화면": f.get("state_label", ""),
                "기준ID": f.get("rule_id", ""),
                "분류": f.get("category", ""),
                "상태": f.get("status", ""),
                "심각도": f.get("severity", ""),
                "내용": f.get("message", ""),
                "상세": f.get("detail", ""),
                "관련근거": _ref_text(f),
                "개선안": fix,
                "캡처": f.get("screenshot_filename", ""),
            }
        )
    cols = [
        "대상유형",
        "화면번호",
        "위치",
        "화면",
        "기준ID",
        "분류",
        "상태",
        "심각도",
        "내용",
        "상세",
        "관련근거",
        "개선안",
        "캡처",
    ]
    if not rows:
        return pd.DataFrame(columns=cols)
    return pd.DataFrame(rows)


def _captures_df(payload: dict[str, Any]) -> pd.DataFrame:
    screen_map = _screen_index_map(payload)
    rows = []
    for s in payload.get("screenshots") or []:
        rows.append(
            {
                "화면번호": _screen_no_label(screen_map, s.get("state_id")),
                "종류": "화면 전체" if s.get("kind") == "state" else "미흡 요소",
                "화면ID": s.get("state_id", ""),
                "화면명": s.get("label", ""),
                "설명": s.get("description", ""),
                "파일": s.get("filename", ""),
                "선택자": s.get("selector", ""),
                "연결finding": s.get("finding_id", ""),
            }
        )
    if not rows:
        return pd.DataFrame(
            columns=[
                "화면번호",
                "종류",
                "화면ID",
                "화면명",
                "설명",
                "파일",
                "선택자",
                "연결finding",
            ]
        )
    return pd.DataFrame(rows)


def build_xlsx_bytes(payload: dict[str, Any], *, embed_images: bool = False) -> bytes:
    findings = payload.get("findings") or []
    coverage = payload.get("coverage") or {}
    stats = payload.get("stats") or {}
    screenshots = payload.get("screenshots") or []
    diff = payload.get("diff") or {}

    summary_rows = [
        {"항목": "진단 대상", "값": payload.get("target_name", "")},
        {"항목": "URL", "값": payload.get("base_url", "")},
        {"항목": "진단 일시", "값": payload.get("scanned_at", "")},
        {"항목": "전체", "값": stats.get("total", 0)},
        {"항목": "통과", "값": stats.get("pass", 0)},
        {"항목": "미흡", "값": stats.get("fail", 0)},
        {"항목": "검토", "값": stats.get("review", 0)},
        {"항목": "미실행", "값": stats.get("not_scanned", 0)},
        {"항목": "해당없음", "값": stats.get("na", 0)},
        {"항목": "캡처 수", "값": len(screenshots)},
        {"항목": "런타임", "값": "가능" if payload.get("runtime_available") else "불가"},
    ]
    if diff:
        summary_rows.extend(
            [
                {"항목": "diff 신규", "값": diff.get("new_count", 0)},
                {"항목": "diff 해소", "값": diff.get("resolved_count", 0)},
                {"항목": "diff 유지", "값": diff.get("unchanged_count", 0)},
            ]
        )

    src_cov = coverage.get("sources") or []
    scr_cov = coverage.get("screens") or []
    screen_map = _screen_index_map(payload)

    manual_rules = []
    for rule in (payload.get("rules") or {}).get("egov") or []:
        if rule.get("automatable") == "manual":
            manual_rules.append(
                {
                    "기준ID": rule.get("id"),
                    "분류": rule.get("category"),
                    "항목": rule.get("title"),
                    "설명": rule.get("description"),
                    "확인": "",
                }
            )
    for rule in (payload.get("rules") or {}).get("krds_uiux") or []:
        if rule.get("automatable") == "manual":
            manual_rules.append(
                {
                    "기준ID": rule.get("id"),
                    "분류": "UI·UX(KRDS)",
                    "항목": rule.get("title"),
                    "설명": rule.get("description"),
                    "확인": "",
                }
            )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="요약", index=False)
        _findings_df(findings, screen_map).to_excel(writer, sheet_name="전체결과", index=False)
        if diff.get("new"):
            _findings_df(diff["new"], screen_map).to_excel(writer, sheet_name="Diff_신규", index=False)
        if diff.get("resolved"):
            _findings_df(diff["resolved"], screen_map).to_excel(
                writer, sheet_name="Diff_해소", index=False
            )
        _captures_df(payload).to_excel(writer, sheet_name="화면캡처", index=False)
        pd.DataFrame(
            [
                {
                    "경로": s.get("path"),
                    "스캔됨": "Y" if s.get("scanned") else "N",
                    "사유": s.get("reason", ""),
                }
                for s in src_cov
                if not s.get("scanned")
            ]
        ).to_excel(writer, sheet_name="미실행_소스", index=False)
        pd.DataFrame(
            [
                {
                    "화면번호": _screen_no_label(screen_map, s.get("state_id")),
                    "화면ID": s.get("state_id"),
                    "화면명": s.get("label"),
                    "설명": s.get("description", ""),
                    "스캔됨": "Y" if s.get("scanned") else "N",
                    "사유": s.get("reason", ""),
                    "캡처파일": next(
                        (
                            sh.get("filename")
                            for sh in screenshots
                            if sh.get("kind") == "state"
                            and sh.get("state_id") == s.get("state_id")
                        ),
                        "",
                    ),
                }
                for s in scr_cov
            ]
        ).to_excel(writer, sheet_name="화면커버리지", index=False)
        pd.DataFrame(manual_rules).to_excel(writer, sheet_name="수동확인", index=False)
        kwcag = (payload.get("rules") or {}).get("kwcag") or []
        egov = (payload.get("rules") or {}).get("egov") or []
        krds = (payload.get("rules") or {}).get("krds_uiux") or []
        pd.DataFrame(kwcag).to_excel(writer, sheet_name="KWCAG22", index=False)
        pd.DataFrame(egov).to_excel(writer, sheet_name="전자정부기준", index=False)
        pd.DataFrame(krds).to_excel(writer, sheet_name="KRDS_UIUX", index=False)

        if embed_images and screenshots:
            ws = writer.sheets["화면캡처"]
            ws.column_dimensions["A"].width = 10
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 18
            ws.column_dimensions["E"].width = 36
            ws.column_dimensions["F"].width = 36
            row = 2
            img_col = len(_captures_df(payload).columns) + 2
            img_letter = get_column_letter(img_col)
            ws.cell(row=1, column=img_col, value="미리보기")
            for shot in screenshots:
                b64 = shot.get("data_base64")
                if not b64:
                    row += 1
                    continue
                try:
                    img_bytes = base64.b64decode(b64)
                    img = XLImage(io.BytesIO(img_bytes))
                    img.width = min(img.width, 480)
                    img.height = min(img.height, 270)
                    ws.add_image(img, f"{img_letter}{row}")
                    ws.row_dimensions[row].height = 200
                except Exception:
                    pass
                row += 1

    buf.seek(0)
    return buf.read()


def build_html_report(
    payload: dict[str, Any],
    *,
    image_mode: Literal["embed", "relative"] = "embed",
) -> str:
    findings = payload.get("findings") or []
    stats = payload.get("stats") or {}
    coverage = payload.get("coverage") or {}
    screenshots = payload.get("screenshots") or []
    shot_map = _screenshot_map(payload)
    screen_map = _screen_index_map(payload)
    diff = payload.get("diff") or {}
    title = f"웹 품질 진단 — {html.escape(str(payload.get('target_name', '')))}"

    def rows_for(items: list[dict], cols: list[str]) -> str:
        if not items:
            return "<tr><td colspan=\"4\">없음</td></tr>"
        out = []
        for item in items:
            out.append(
                "<tr>"
                + "".join(f"<td>{html.escape(str(item.get(c, '')))}</td>" for c in cols)
                + "</tr>"
            )
        return "\n".join(out)

    fail_rows = [
        f
        for f in findings
        if f.get("status") in ("fail", "review", "not_scanned")
    ][:200]

    capture_sections = []
    state_shots = [s for s in screenshots if s.get("kind") == "state"]
    for shot in state_shots:
        src = _image_src(shot, image_mode)
        if not src:
            continue
        screen_no = _screen_no_label(screen_map, shot.get("state_id"))
        no_prefix = (
            f'<span class="screen-no">{html.escape(screen_no)}</span> '
            if screen_no
            else ""
        )
        capture_sections.append(
            f"""<section class="capture-card">
<h3>{no_prefix}{html.escape(str(shot.get('label', '')))} <span class="muted">({html.escape(str(shot.get('state_id', '')))})</span></h3>
<p class="capture-desc">{html.escape(str(shot.get('description', '')))}</p>
<img src="{src}" alt="{html.escape(str(shot.get('label', '')))}" loading="lazy"/>
</section>"""
        )

    element_shots = [s for s in screenshots if s.get("kind") == "element"]
    element_blocks = []
    for shot in element_shots:
        src = _image_src(shot, image_mode)
        if not src:
            continue
        screen_no = _screen_no_label(screen_map, shot.get("state_id"))
        no_prefix = (
            f'<span class="screen-no">{html.escape(screen_no)}</span> '
            if screen_no
            else ""
        )
        element_blocks.append(
            f"""<div class="element-shot">
<p>{no_prefix}<strong>{html.escape(str(shot.get('label', '')))}</strong> — {html.escape(str(shot.get('description', '')))}</p>
<p class="muted">{html.escape(str(shot.get('selector', '')))}</p>
<img src="{src}" alt="미흡 요소" loading="lazy"/>
</div>"""
        )

    fail_table_rows = []
    for x in fail_rows:
        shot_html = ""
        sid = x.get("screenshot_id")
        if sid and sid in shot_map:
            src = _image_src(shot_map[sid], image_mode)
            if src:
                shot_html = f'<br/><img class="inline-shot" src="{src}" alt="캡처"/>'
        fix_cell = html.escape(str(x.get("fix", "")))
        if x.get("fix_url"):
            fix_cell += (
                f'<br/><a href="{html.escape(str(x.get("fix_url")))}" '
                f'target="_blank" rel="noopener">참고 문서</a>'
            )
        fail_table_rows.append(
            f"<tr><td>{html.escape(str(x.get('target','')))}</td>"
            f"<td>{_finding_screen_no_html(x, screen_map)}</td>"
            f"<td>{html.escape(str(x.get('location','')))}</td>"
            f"<td>{html.escape(str(x.get('rule_id','')))}</td>"
            f"<td class='{html.escape(str(x.get('status','fail')))}'>{html.escape(str(x.get('status','')))}</td>"
            f"<td>{html.escape(str(x.get('message','')))}{shot_html}</td>"
            f"<td class='ref'>{_ref_cell_html(x)}</td>"
            f"<td class='fix'>{fix_cell}</td></tr>"
        )

    not_scanned_screens = [
        {
            "screen_no": _screen_no_label(screen_map, s.get("state_id")),
            "state_id": s.get("state_id", ""),
            "label": s.get("label", ""),
            "description": s.get("description", ""),
            "reason": s.get("reason", ""),
        }
        for s in (coverage.get("screens") or [])
        if not s.get("scanned")
    ]

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8"/>
<title>{title}</title>
<style>
body {{ font-family: "Malgun Gothic", sans-serif; margin: 24px; color: #111; }}
h1,h2,h3 {{ margin-top: 1.2em; }}
.stats {{ display: grid; grid-template-columns: repeat(auto-fill,minmax(120px,1fr)); gap: 8px; }}
.stat {{ border: 1px solid #ccc; padding: 8px; border-radius: 6px; }}
.capture-grid {{ display: grid; grid-template-columns: repeat(auto-fill,minmax(360px,1fr)); gap: 16px; }}
.capture-card {{ border: 1px solid #ddd; border-radius: 8px; padding: 12px; }}
.capture-card h3 {{ margin: 0 0 4px; }}
.capture-card .capture-desc {{ margin: 0 0 8px; }}
.screen-no {{ display: inline-block; min-width: 2em; font-weight: 700; color: #0b57d0; background: #e8f0fe; border: 1px solid #c6dafc; border-radius: 4px; padding: 1px 6px; margin-right: 6px; font-size: 12px; vertical-align: middle; }}
.capture-card img, .element-shot img {{ max-width: 100%; height: auto; border: 1px solid #eee; margin-top: 8px; display: block; }}
.element-shot {{ border-left: 3px solid #b00020; padding-left: 12px; margin: 12px 0; }}
.inline-shot {{ max-width: 320px; display: block; margin-top: 6px; }}
.muted {{ color: #666; font-size: 13px; }}
table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
th, td {{ border: 1px solid #ddd; padding: 6px 8px; text-align: left; vertical-align: top; }}
th {{ background: #f4f4f4; }}
.fail {{ color: #b00020; }}
.review {{ color: #b8860b; }}
.ref-link {{ color: #0b57d0; text-decoration: none; }}
.ref-link:hover {{ text-decoration: underline; }}
.ref-krds {{ font-weight: 600; }}
.ref-extra {{ margin-top: 4px; font-size: 12px; }}
.ref-fallback {{ color: #555; }}
@media print {{ body {{ margin: 12px; }} .capture-card {{ break-inside: avoid; }} }}
</style>
</head>
<body>
<h1>{title}</h1>
<p>URL: {html.escape(str(payload.get('base_url','')))}<br/>
진단 일시: {html.escape(str(payload.get('scanned_at','')))}</p>
<div class="stats">
  <div class="stat"><strong>전체</strong><br/>{stats.get('total',0)}</div>
  <div class="stat"><strong>통과</strong><br/>{stats.get('pass',0)}</div>
  <div class="stat"><strong>미흡</strong><br/>{stats.get('fail',0)}</div>
  <div class="stat"><strong>검토</strong><br/>{stats.get('review',0)}</div>
  <div class="stat"><strong>미실행</strong><br/>{stats.get('not_scanned',0)}</div>
  <div class="stat"><strong>캡처</strong><br/>{len(screenshots)}</div>
</div>
{"<h2>이전 대비</h2><p>신규 " + str(diff.get('new_count',0)) + " · 해소 " + str(diff.get('resolved_count',0)) + " · 유지 " + str(diff.get('unchanged_count',0)) + "</p>" if diff else ""}
<h2>화면 캡처</h2>
<div class="capture-grid">
{"".join(capture_sections) if capture_sections else "<p>화면 캡처 없음 (런타임 진단 미실행)</p>"}
</div>
<h2>미흡 요소 캡처</h2>
{"".join(element_blocks) if element_blocks else "<p>미흡 요소 캡처 없음</p>"}
<h2>미실행 소스</h2>
<table><thead><tr><th>경로</th><th>사유</th></tr></thead><tbody>
{rows_for([s for s in (coverage.get('sources') or []) if not s.get('scanned')], ['path','reason'])}
</tbody></table>
<h2>미실행 화면</h2>
<table><thead><tr><th>화면번호</th><th>화면ID</th><th>화면명</th><th>설명</th><th>사유</th></tr></thead><tbody>
{rows_for(not_scanned_screens, ['screen_no','state_id','label','description','reason'])}
</tbody></table>
<h2>주요 미흡·미실행 항목</h2>
<table>
<thead><tr><th>유형</th><th>화면번호</th><th>위치</th><th>기준</th><th>상태</th><th>내용</th><th>관련근거</th><th>개선안</th></tr></thead>
<tbody>
{"".join(fail_table_rows)}
</tbody></table>
</body>
</html>"""


def build_zip_bytes(payload: dict[str, Any]) -> bytes:
    html_body = build_html_report(payload, image_mode="relative")
    xlsx_body = build_xlsx_bytes(payload, embed_images=True)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("report.html", html_body.encode("utf-8"))
        zf.writestr("report.xlsx", xlsx_body)
        for shot in payload.get("screenshots") or []:
            b64 = shot.get("data_base64")
            name = shot.get("filename")
            if b64 and name:
                zf.writestr(name, base64.b64decode(b64))
    buf.seek(0)
    return buf.read()
